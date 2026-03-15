import pandas as pd, os
from openpyxl.styles import PatternFill, Font, Alignment
from database import get_businesses_by_session

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "business_data_export.xlsx")

COL_ORDER = [
    "name", "category", "address", "city", "state", "country",
    "phone", "website", "rating", "review_count", "price_level",
    "business_status", "opening_hours", "latitude", "longitude",
    "maps_url", "date_scraped"
]

COL_WIDTHS = {
    "name": 28, "category": 14, "address": 38, "city": 14,
    "state": 14, "country": 12, "phone": 18, "website": 32,
    "rating": 8, "review_count": 12, "price_level": 10,
    "business_status": 18, "opening_hours": 48,
    "latitude": 12, "longitude": 12, "maps_url": 36, "date_scraped": 20,
}

def export_to_excel(session_id, business_type="businesses", city=""):
    records = get_businesses_by_session(session_id)
    if not records:
        return None

    df = pd.DataFrame(records)
    present = [c for c in COL_ORDER if c in df.columns]
    df = df[present]
    df.columns = [c.replace("_", " ").title() for c in present]

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Businesses")
        ws = writer.sheets["Businesses"]

        # Header styling using the brand palette
        header_fill = PatternFill("solid", fgColor="5C164E")
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = Font(color="E2FCEF", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        for i, col_name in enumerate(present, start=1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_name, 18)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Alternate row shading
        light_fill = PatternFill("solid", fgColor="F9F5F3")
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = light_fill

    return OUTPUT_PATH
